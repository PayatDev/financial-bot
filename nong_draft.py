import os
import json
import tempfile
from datetime import datetime

import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build

from drive_service import create_folder, upload_file_to_folder
from nong_doc import run as doc_run

DEV_MODE = True
SHEET_ID  = os.environ.get("GOOGLE_SHEET_ID")
SCOPES    = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
MODEL = os.environ.get("ANTHROPIC_MODEL", "claude-sonnet-4-5-20250929")

NAVY  = "1E3A5F"
MGRY  = "E8ECF0"
WHITE = "FFFFFF"

FIELD_LABELS = [
    ("nickname","ชื่อเล่น"),("age","อายุ"),("gender","เพศ"),
    ("occupation","อาชีพ"),("health","สุขภาพ"),("email","อีเมล"),
    ("income_self","รายได้"),("hobbies_and_risks","งานอดิเรก"),
    ("spouse_nickname","คู่สมรส"),("spouse_age","อายุคู่สมรส"),
    ("spouse_occupation","อาชีพคู่สมรส"),("spouse_income","รายได้คู่สมรส"),
    ("spouse_health","สุขภาพคู่สมรส"),("spouse_status","สถานะ"),
    ("children","ลูก"),("children_outside_marriage","บุตรนอกสมรส"),
    ("assets_cash","เงินสด"),("assets_property","อสังหาริมทรัพย์"),
    ("assets_investment","การลงทุน"),("assets_crypto_wallet","คริปโต"),
    ("assets_insurance_savings","ประกันสะสม"),("assets_digital","ทรัพย์สินดิจิทัล"),
    ("assets_business","กิจการ"),("assets_valuables","ของมีค่า"),
    ("debt","หนี้สิน"),("guarantor","ค้ำประกัน"),
    ("insurance_life","ประกันชีวิต"),("insurance_health","ประกันสุขภาพ"),
    ("insurance_group","ประกันกลุ่ม"),("welfare","สวัสดิการ"),
    ("funeral_wishes","ความปรารถนางานศพ"),("emergency_cash_90days","เงินฉุกเฉิน 90 วัน"),
    ("estate_admin_cost","ต้นทุนจัดการมรดก"),("asset_distribution","แผนแบ่งทรัพย์"),
    ("debt_responsibility","หนี้ใครรับผิดชอบ"),("business_succession","แผนกิจการ"),
    ("urgent_manager","ผู้จัดการฉุกเฉิน"),("estate_executor","ผู้จัดการมรดก"),
    ("financial_poa","Financial POA"),("living_will","Living Will"),
    ("surviving_spouse_plan","แผนคู่สมรสที่รอดชีวิต"),
    ("guardian_primary","Guardian หลัก"),("guardian_backup","Guardian สำรอง"),
    ("money_guardian_primary","Money Guardian หลัก"),("money_guardian_backup","Money Guardian สำรอง"),
    ("documents_location","ที่อยู่เอกสาร"),
    ("letter_to_children","จดหมายถึงลูก"),("letter_to_spouse","จดหมายถึงคู่สมรส"),
    ("fullname_self","ชื่อ-นามสกุลจริงเจ้าของแผน"),("id_self","เลขบัตรประชาชนเจ้าของแผน"),
    ("address_self","ที่อยู่ปัจจุบัน"),
    ("fullname_spouse","ชื่อ-นามสกุลจริงคู่สมรส"),("id_spouse","เลขบัตรประชาชนคู่สมรส"),
    ("fullname_children","ชื่อ-นามสกุลจริงลูก (ทุกคน)"),
    ("fullname_executor","ชื่อ-นามสกุลจริงผู้จัดการมรดก"),("id_executor","เลขบัตรประชาชนผู้จัดการมรดก"),
    ("fullname_executor_backup","ชื่อ-นามสกุลจริงผู้จัดการมรดกสำรอง"),("id_executor_backup","เลขบัตรประชาชนผู้จัดการมรดกสำรอง"),
    ("fullname_guardian_primary","ชื่อ-นามสกุลจริงผู้ปกครองหลัก"),("id_guardian_primary","เลขบัตรประชาชนผู้ปกครองหลัก"),
    ("fullname_guardian_backup","ชื่อ-นามสกุลจริงผู้ปกครองสำรอง"),("id_guardian_backup","เลขบัตรประชาชนผู้ปกครองสำรอง"),
    ("fullname_money_guardian_primary","ชื่อ-นามสกุลจริงผู้ดูแลเงินหลัก"),("id_money_guardian_primary","เลขบัตรประชาชนผู้ดูแลเงินหลัก"),
    ("fullname_money_guardian_backup","ชื่อ-นามสกุลจริงผู้ดูแลเงินสำรอง"),("id_money_guardian_backup","เลขบัตรประชาชนผู้ดูแลเงินสำรอง"),
]

# ── helpers ─────────────────────────────────────────────────────────────
def fill(h):    return PatternFill("solid", start_color=h, end_color=h)
def fnt(bold=False, size=10, color="222222", italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)
def thin(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def cal(h="left"): return Alignment(horizontal=h, vertical="center", wrap_text=True)
def put(ws, row, col, val="", bg=WHITE, bold=False, color="222222",
        size=10, align="left", bc="CCCCCC"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill(bg); c.font = fnt(bold, size, color)
    c.alignment = cal("center" if align == "center" else "left")
    c.border = thin(bc); return c


# ── Google Sheets ───────────────────────────────────────────────────────
def get_data_from_sheets():
    creds = Credentials.from_service_account_info(
        json.loads(os.environ.get("GOOGLE_CREDENTIALS_JSON")), scopes=SCOPES)
    svc = build("sheets", "v4", credentials=creds)
    result = svc.spreadsheets().values().get(
        spreadsheetId=SHEET_ID, range="Sheet1!A1:BZ2").execute()
    values = result.get("values", [])
    if len(values) < 2: return {}
    headers = values[0]; row = values[1]
    row += [""] * (len(headers) - len(row))
    return dict(zip(headers, row))


# ── Claude ──────────────────────────────────────────────────────────────
def call_claude(data: dict) -> str:
    prompt = f"""คุณคือผู้ช่วยของคุณพยัต นักวางแผนการเงินและกฎหมาย

เล่าเรื่องราวของลูกค้าต่อเนื่อง 3-5 ย่อหน้า ภาษาไทยธรรมดา อ่านเข้าใจง่าย
ครอบคลุมทุก field ที่มีข้อมูล ข้ามข้อมูลที่เป็น ไม่มี หรือ ไม่ได้ระบุ
ไม่ต้อง output อะไรนอกจากเรื่องราว ไม่ต้องมี header ไม่ต้องมี JSON

ข้อมูลลูกค้า:
{json.dumps(data, ensure_ascii=False, indent=2)}"""

    client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))
    print("กำลังสร้างเรื่องราว...")
    r = client.messages.create(
        model=MODEL, max_tokens=3000,
        messages=[{"role": "user", "content": prompt}]
    )
    story = r.content[0].text.strip()
    print("✅ story พร้อม")
    return story


# ── build workbook ──────────────────────────────────────────────────────
def build_workbook(data: dict, story: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "เรื่องราวลูกค้า"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 54

    # header
    ws.row_dimensions[1].height = 32
    c = ws.cell(row=1, column=1,
        value=f"{data.get('nickname','')} | {data.get('occupation','')} | {data.get('age','')} ปี")
    c.fill = fill(NAVY); c.font = fnt(bold=True, size=13, color=WHITE)
    c.alignment = cal(); c.border = thin(NAVY)
    ws.merge_cells("A1:B1")

    row = 2

    def header_row(text):
        nonlocal row
        ws.row_dimensions[row].height = 18
        c = ws.cell(row=row, column=1, value=text)
        c.fill = fill(MGRY); c.font = fnt(bold=True, size=10, color=NAVY)
        c.alignment = cal(); c.border = thin(MGRY)
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

    # Block A — เรื่องราว
    header_row("เรื่องราวลูกค้า")
    for para in story.split("\n"):
        para = para.strip().lstrip("#").lstrip("*").strip()
        if not para: continue
        ws.row_dimensions[row].height = 60
        c = ws.cell(row=row, column=1, value=para)
        c.fill = fill("F5F7FA"); c.font = fnt(size=10)
        c.alignment = cal(); c.border = thin()
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

    # Block B — ข้อมูลรายช่อง
    header_row("ข้อมูลรายช่อง")
    for key, label in FIELD_LABELS:
        val = data.get(key, "")
        if not val or val in ("ไม่มี", "ไม่ได้ระบุ", ""): continue
        ws.row_dimensions[row].height = 30
        put(ws, row, 1, label, bg="F5F7FA", bold=True, color="444444", size=9)
        put(ws, row, 2, val, size=10)
        row += 1

    ws.freeze_panes = "A2"

    nickname = data.get("nickname", "ลูกค้า")
    date_str = datetime.now().strftime("%d%m%Y")
    filename = f"เรื่องราว_คุณ{nickname}.xlsx"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    return tmp.name, filename


# ── entry point ─────────────────────────────────────────────────────────
def run(data=None):
    if DEV_MODE and data is None:
        print("DEV MODE — ดึงข้อมูลจาก Sheets")
        data = get_data_from_sheets()

    print(f"ข้อมูล: {data.get('nickname')} | {data.get('occupation')} | {data.get('age')} ปี")

    story = call_claude(data)

    nickname    = data.get("nickname", "ลูกค้า")
    date_str    = datetime.now().strftime("%d%m%Y")
    folder_name = f"{date_str}_แผนครอบครัว_คุณ{nickname}"
    folder_id   = create_folder(folder_name)

    local_path, filename = build_workbook(data, story)
    print(f"สร้างไฟล์: {filename}")
    upload_file_to_folder(local_path, filename, folder_id)
    if os.path.exists(local_path): os.unlink(local_path)

    print("✅ xlsx พร้อม → เริ่มสร้างเอกสาร")
    doc_run(folder_name, folder_id)
