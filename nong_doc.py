import os
import json
import tempfile
import io
from datetime import datetime

import anthropic
from openpyxl import load_workbook
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request

from drive_service import get_drive_service, upload_file_to_folder
from email_service import send_line_message

from cover_builder import generate_issue_content, build_cover
from will_builder import build_will
from poa_builder import build_poa
from living_will_builder import build_living_will
from emergency_guide_builder import build_emergency_guide


def find_folder_id(folder_name: str) -> str:
    service = get_drive_service()
    result = service.files().list(
        q=f"name='{folder_name}' and mimeType='application/vnd.google-apps.folder'",
        fields="files(id, name)"
    ).execute()
    files = result.get("files", [])
    if not files:
        raise ValueError(f"ไม่พบ folder: {folder_name}")
    return files[0]["id"]


def download_xlsx(folder_id: str) -> str:
    service = get_drive_service()
    result = service.files().list(
        q=f"'{folder_id}' in parents and name contains '.xlsx'",
        fields="files(id, name)"
    ).execute()
    files = result.get("files", [])
    files = [f for f in files if not f["name"].startswith("~$")]
    if not files:
        raise ValueError("ไม่พบ xlsx ใน folder")
    file_id = files[0]["id"]
    request = service.files().get_media(fileId=file_id)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    downloader = MediaIoBaseDownload(tmp, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    tmp.close()
    print(f"✅ Downloaded xlsx: {files[0]['name']}")
    return tmp.name

def read_xlsx(local_path: str) -> dict:
    wb = load_workbook(local_path, data_only=True)
    ws1 = wb.worksheets[0]  # ดึง sheet แรกเสมอ
    client_data = {}
    for row in ws1.iter_rows(values_only=True):
        if row[0] and row[1] and row[0] not in ("เรื่องราวลูกค้า", "ข้อมูลรายช่อง"):
            if isinstance(row[0], str) and len(row[0]) < 40:
                client_data[row[0]] = row[1]
    return client_data


def run(folder_name: str, folder_id: str = None):
    print(f"น้องดอค เริ่มทำงาน: {folder_name}")

    if folder_id is None:
        folder_id = find_folder_id(folder_name)
    print(f"✅ Found folder: {folder_id}")

    xlsx_path = download_xlsx(folder_id)
    client_data = read_xlsx(xlsx_path)
    if os.path.exists(xlsx_path):
        os.unlink(xlsx_path)
    print(f"✅ อ่าน xlsx เรียบร้อย: {len(client_data)} fields")

    print("กำลัง generate เนื้อหา...")
    generated = generate_issue_content(client_data)
    print("✅ Generate เสร็จ")

    nickname = client_data.get("ชื่อเล่น", folder_name.split("_")[0])

    cover_path = build_cover(client_data, [], generated, folder_name)
    cover_filename = f"1_แผนครอบครัว_คุณ{nickname}.docx"
    upload_file_to_folder(cover_path, cover_filename, folder_id)
    if os.path.exists(cover_path): os.unlink(cover_path)

    will_path = build_will(client_data)
    will_filename = f"2_พินัยกรรม_คุณ{nickname}.docx"
    upload_file_to_folder(will_path, will_filename, folder_id)
    if os.path.exists(will_path): os.unlink(will_path)

    poa_path = build_poa(client_data)
    poa_filename = f"3_หนังสือมอบอำนาจ_คุณ{nickname}.docx"
    upload_file_to_folder(poa_path, poa_filename, folder_id)
    if os.path.exists(poa_path): os.unlink(poa_path)

    lw_path = build_living_will(client_data)
    lw_filename = f"4_หนังสือแสดงเจตนาการยื้อชีวิต_คุณ{nickname}.docx"
    upload_file_to_folder(lw_path, lw_filename, folder_id)
    if os.path.exists(lw_path): os.unlink(lw_path)

    guide_path = build_emergency_guide(client_data)
    guide_filename = f"5_คู่มือฉุกเฉิน_คุณ{nickname}.docx"
    upload_file_to_folder(guide_path, guide_filename, folder_id)
    if os.path.exists(guide_path): os.unlink(guide_path)

    send_line_message(
        f"✅ เอกสารพร้อมแล้วครับ\n\n"
        f"ชื่อ: คุณ{nickname}\n"
        f"อาชีพ: {client_data.get('อาชีพ', '')}\n"
        f"อายุ: {client_data.get('อายุ', '')} ปี\n"
        f"📁 {folder_name}\n\n"
        f"เอกสาร 5 ชิ้นพร้อมใน Drive แล้วครับ\n"
        f"โปรดตรวจและแก้ไขก่อนส่งลูกค้าทาง\n"
        f"📧 {client_data.get('อีเมล', '[ยังไม่มีอีเมล]')}"
    )
    print("เสร็จสิ้น")
